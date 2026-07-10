"""
nova/export.py

Phase 3: Write extracted data to a real spreadsheet file.

Takes the list[dict] output of extraction.py and writes it to a .xlsx
file you can open in Excel/Google Sheets - the actual "save the top 10
in a spreadsheet" part of the original Nova vision.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def save_to_xlsx(data: list[dict], filepath: str, sheet_title: str = "Nova Results"):
    """
    Write a list of dicts to an .xlsx file. Column headers are taken from
    the union of all keys across all dicts, in first-seen order.
    """
    if not data:
        print("[export] No data to save - skipping file creation.")
        return None

    # collect columns in first-seen order across all rows
    columns: list[str] = []
    for row in data:
        for key in row.keys():
            if key not in columns:
                columns.append(key)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]  # Excel sheet name limit

    # header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill

    # data rows
    for row_idx, row in enumerate(data, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col_name))

    # auto-width columns (rough heuristic, good enough for readability)
    for col_idx, col_name in enumerate(columns, start=1):
        max_len = max(
            [len(str(col_name))] + [len(str(row.get(col_name, ""))) for row in data]
        )
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 50)

    wb.save(filepath)
    print(f"[export] Saved {len(data)} rows to {filepath}")
    return filepath